#!/usr/bin/env python3
"""Update blue-highlighted cells in the costing Excel file with new prices.

Usage:
    python tools/update_costing_file.py <excel_path> [--output <output_path>] [options]

Raipur tab options:
    --pallet-dri <value>       Pallet DRI price (INR/MT)
    --pig-iron <value>         Pig Iron price (INR/MT)
    --scrap-raipur <value>     Scrap price DAP Raipur (INR/MT)
    --silico-mn <value>        Silico Manganese price (INR/kg)
    --iron-ore-dri <value>     Iron Ore DRI price (INR/MT)
    --report-date <YYYY-MM-DD> Report date
    --billet-raipur <value>    Market price of Billet Raipur (INR/MT)
    --tmt-raipur <value>       Market price of TMT Raipur (INR/MT)

NCR tab options:
    --scrap-mandi <value>      Scrap price DAP Mandi (INR/MT, before -500 adj)
    --billet-mandi <value>     Billet price Mandi Gobindgarh (INR/MT, before -500 adj)
    --tmt-ncr <value>          Market price of TMT NCR (INR/MT)

Exit codes:
    0 — Success
    1 — Error
"""

import argparse
import os
import subprocess
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font

from format_output import format_change_log, format_ncr, format_raipur

# Items tracked in the change log (fixed order)
LOG_ITEMS = [
    "Pallet DRI",
    "Pig Iron",
    "Scrap",
    "Silico Manganese",
    "Iron Ore DRI",
    "Date",
    "Market price Billet",
    "Nett Margin Billet",
    "Market price TMT",
    "Margin TMT",
]


def _verify_no_formulas(filepath):
    """Scan the saved Excel file and fail loudly if any formula strings remain."""
    wb = openpyxl.load_workbook(filepath)
    errors = []
    for ws in wb:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    errors.append(f"  {ws.title}!{cell.coordinate}: {cell.value}")
    if errors:
        print("FATAL: Formulas found in output file (would cause Excel errors):", file=sys.stderr)
        for e in errors:
            print(e, file=sys.stderr)
        sys.exit(1)
    print(f"Verified: 0 formulas in output (all values computed)", file=sys.stderr)


def _auto_push(output_path, log_path, report_date):
    """Auto-commit and push output files to main."""
    try:
        # Stage output files
        files_to_add = [output_path, log_path]
        files_to_add = [f for f in files_to_add if os.path.exists(f)]
        subprocess.run(["git", "add"] + files_to_add, check=True, capture_output=True)

        # Commit
        msg = f"Update costing output for {report_date}"
        subprocess.run(["git", "commit", "-m", msg], check=True, capture_output=True)

        # Push to main with retry
        for attempt in range(4):
            result = subprocess.run(
                ["git", "push", "origin", "main"],
                capture_output=True, text=True,
            )
            if result.returncode == 0:
                print(f"Pushed to main successfully", file=sys.stderr)
                return
            if attempt < 3:
                import time
                wait = 2 ** (attempt + 1)
                print(f"Push failed, retrying in {wait}s...", file=sys.stderr)
                time.sleep(wait)
        print(f"WARNING: Push failed after 4 attempts. Commit saved locally.", file=sys.stderr)
    except subprocess.CalledProcessError as e:
        # Nothing to commit is OK (e.g., re-run for same date)
        if b"nothing to commit" in (e.stderr or b""):
            print("No new changes to push (same data already committed)", file=sys.stderr)
        else:
            print(f"WARNING: Auto-push failed: {e.stderr}", file=sys.stderr)


def _cell_val(ws, ref, fallback=0):
    """Read a cell value, returning fallback if it's a formula or None."""
    v = ws[ref].value
    if v is None or (isinstance(v, str) and v.startswith("=")):
        return fallback
    return float(v)


def compute_margins(wb, args):
    """Compute Nett Margin (Billet) and Margin (TMT) for both markets.

    Since openpyxl can't evaluate formulas, we replicate the Excel
    calculation chain using the cell constants (consumption %, recovery
    ratios, operating costs) already in the workbook.
    """
    margins = {}

    # --- Raipur ---
    ws = wb["Raipur"]
    # Billet cost components (I7:I14)
    # I = (E * C/100) / G  for percentage items, or (E * C) / G for kg items
    i7 = (_cell_val(ws, "E7") * _cell_val(ws, "C7") / 100) / _cell_val(ws, "G7", 1)
    i8 = (_cell_val(ws, "E8") * _cell_val(ws, "C8") / 100) / _cell_val(ws, "G8", 1)
    i9 = (_cell_val(ws, "E9") * _cell_val(ws, "C9") / 100) / _cell_val(ws, "G9", 1)
    i10 = (_cell_val(ws, "E10") * _cell_val(ws, "C10")) / _cell_val(ws, "G10", 1)
    i11 = (_cell_val(ws, "E11") * _cell_val(ws, "C11") / 100) / _cell_val(ws, "G11", 1)
    i12 = _cell_val(ws, "E12") * _cell_val(ws, "C12")
    i13 = _cell_val(ws, "I13")  # hardcoded in Excel
    i14 = _cell_val(ws, "I14")  # hardcoded in Excel
    raipur_billet_cost = i7 + i8 + i9 + i10 + i11 + i12 + i13 + i14
    raipur_billet_mkt = _cell_val(ws, "I16")
    margins["raipur_billet"] = round(raipur_billet_mkt - raipur_billet_cost, 2)

    # Rolling mill cost (F22:F27 summed = F28, then F33 = F28 + I15 + F32)
    f22 = raipur_billet_cost * _cell_val(ws, "C22") / 100  # scale loss
    f23 = _cell_val(ws, "E23") * _cell_val(ws, "C23")  # power
    f24 = _cell_val(ws, "E24") * _cell_val(ws, "C24")  # stores
    f25 = _cell_val(ws, "E25") * _cell_val(ws, "C25")  # manpower
    f26 = _cell_val(ws, "E26") * _cell_val(ws, "C26")  # fuel
    f27_base = raipur_billet_cost + f22 + f23 + f24 + f25 + f26
    f27 = f27_base * _cell_val(ws, "C27") / 100  # cutting loss
    f28 = f22 + f23 + f24 + f25 + f26 + f27  # rolling cost
    f32 = _cell_val(ws, "F32")  # interest + depreciation
    raipur_total_cost = f28 + raipur_billet_cost + f32
    raipur_tmt_mkt = _cell_val(ws, "F34")
    margins["raipur_tmt"] = round(raipur_tmt_mkt - raipur_total_cost, 2)

    # --- NCR ---
    ws_n = wb["NCR"]
    # NCR rates: D7=Raipur!E7+3100, D8=Raipur!E8+3100, D9=scrap-500,
    # D10=Raipur!E10, D11=Raipur!E11+3100
    d7 = _cell_val(ws, "E7") + 3100
    d8 = _cell_val(ws, "E8") + 3100
    d9 = (args.scrap_mandi - 500) if args.scrap_mandi is not None else _cell_val(ws_n, "D9")
    d10 = _cell_val(ws, "E10")  # same as Raipur
    d11 = _cell_val(ws, "E11") + 3100

    h7 = (d7 * _cell_val(ws_n, "B7") / 100) / _cell_val(ws_n, "F7", 1)
    h8 = (d8 * _cell_val(ws_n, "B8") / 100) / _cell_val(ws_n, "F8", 1)
    h9 = (d9 * _cell_val(ws_n, "B9") / 100) / _cell_val(ws_n, "F9", 1)
    h10 = (d10 * _cell_val(ws_n, "B10")) / _cell_val(ws_n, "F10", 1)
    h11 = (d11 * _cell_val(ws_n, "B11") / 100) / _cell_val(ws_n, "F11", 1) if _cell_val(ws_n, "B11") != 0 else 0
    h12 = _cell_val(ws_n, "D12") * _cell_val(ws_n, "B12")
    h13 = _cell_val(ws_n, "D13") * _cell_val(ws_n, "B13")
    h14 = _cell_val(ws_n, "D14") * _cell_val(ws_n, "B14")
    ncr_billet_cost = h7 + h8 + h9 + h10 + h11 + h12 + h13 + h14
    ncr_billet_mkt = (args.billet_mandi - 500) if args.billet_mandi is not None else _cell_val(ws_n, "H16")
    margins["ncr_billet"] = round(ncr_billet_mkt - ncr_billet_cost, 2)

    # NCR rolling mill (same structure, reads from NCR tab)
    n_f22 = ncr_billet_cost * _cell_val(ws_n, "C22") / 100
    n_f23 = _cell_val(ws_n, "E23") * _cell_val(ws_n, "C23")
    n_f24 = _cell_val(ws_n, "E24") * _cell_val(ws_n, "C24")
    n_f25 = _cell_val(ws_n, "E25") * _cell_val(ws_n, "C25")
    n_f26 = _cell_val(ws_n, "E26") * _cell_val(ws_n, "C26")
    n_f27_base = ncr_billet_cost + n_f22 + n_f23 + n_f24 + n_f25 + n_f26
    n_f27 = n_f27_base * _cell_val(ws_n, "C27") / 100
    n_f28 = n_f22 + n_f23 + n_f24 + n_f25 + n_f26 + n_f27
    n_f32 = _cell_val(ws_n, "F32")
    ncr_total_cost = n_f28 + ncr_billet_cost + n_f32
    ncr_tmt_mkt = _cell_val(ws_n, "F34")
    margins["ncr_tmt"] = round(ncr_tmt_mkt - ncr_total_cost, 2)

    return margins


def update_change_log(args, margins):
    """Create or update the cumulative change log at output/change_log.xlsx."""
    log_path = os.path.join("output", "change_log.xlsx")
    date_str = args.report_date

    # Collect values (order must match LOG_ITEMS)
    raipur_values = [
        args.pallet_dri,
        args.pig_iron,
        args.scrap_raipur,
        args.silico_mn,
        args.iron_ore_dri,
        args.report_date,
        args.billet_raipur,
        margins.get("raipur_billet"),
        args.tmt_raipur,
        margins.get("raipur_tmt"),
    ]

    ncr_values = [
        "-",  # Pallet DRI — auto-calculated from Raipur
        "-",  # Pig Iron — auto-calculated from Raipur
        (args.scrap_mandi - 500) if args.scrap_mandi is not None else None,
        "-",  # Silico Mn — same as Raipur via formula
        "-",  # Iron Ore DRI — auto-calculated from Raipur
        args.report_date,
        (args.billet_mandi - 500) if args.billet_mandi is not None else None,
        margins.get("ncr_billet"),
        args.tmt_ncr,
        margins.get("ncr_tmt"),
    ]

    # Load or create workbook
    if os.path.exists(log_path):
        log_wb = openpyxl.load_workbook(log_path)
        ws = log_wb.active
    else:
        log_wb = openpyxl.Workbook()
        ws = log_wb.active
        ws.title = "Change Log"
        # Write item labels
        ws.cell(row=1, column=1, value="Item").font = Font(bold=True)
        for i, item in enumerate(LOG_ITEMS):
            ws.cell(row=i + 3, column=1, value=item)

    # Find existing date column or next available slot
    raipur_col = None
    col = 2
    while ws.cell(row=1, column=col).value is not None:
        if str(ws.cell(row=1, column=col).value) == date_str:
            raipur_col = col
            break
        col += 2
    if raipur_col is None:
        raipur_col = col
    ncr_col = raipur_col + 1

    # Write headers
    bold_center = Font(bold=True)
    center = Alignment(horizontal="center")

    ws.cell(row=1, column=raipur_col, value=date_str).font = bold_center
    ws.cell(row=1, column=raipur_col).alignment = center
    ws.cell(row=2, column=raipur_col, value="Raipur").font = bold_center
    ws.cell(row=2, column=raipur_col).alignment = center
    ws.cell(row=2, column=ncr_col, value="NCR").font = bold_center
    ws.cell(row=2, column=ncr_col).alignment = center

    # Write data (items start at row 3)
    for i, (r_val, n_val) in enumerate(zip(raipur_values, ncr_values)):
        row = i + 3
        if r_val is not None:
            ws.cell(row=row, column=raipur_col, value=r_val)
        if n_val is not None:
            ws.cell(row=row, column=ncr_col, value=n_val)

    # Auto-adjust column widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 8)

    os.makedirs("output", exist_ok=True)
    log_wb.save(log_path)
    print(f"Change log updated: {log_path}", file=sys.stderr)


def _compute_tab(ws, rate_cells, is_ncr=False):
    """Compute all formula cells in a tab and write values directly.

    rate_cells is a dict of cell->value for the rate/input cells that were updated.
    Reads constants (consumption, recovery, operating costs) from the sheet.
    """
    # Column mapping: Raipur uses cols B-I, NCR uses cols A-H
    # Rate col: E (Raipur), D (NCR)
    # Cost col: F (both)
    # Recovery col: G (Raipur), F (NCR)
    # Effective col: H (Raipur), G (NCR)
    # Total col: I (Raipur), H (NCR)
    if is_ncr:
        r_col, f_col, g_col, h_col, t_col = "D", "E", "F", "G", "H"
    else:
        r_col, f_col, g_col, h_col, t_col = "E", "F", "G", "H", "I"

    # --- Section 1: Raw material rows (7-11) ---
    total_cost = 0
    for row in range(7, 12):
        rate = _cell_val(ws, f"{r_col}{row}")
        cons = _cell_val(ws, f"{'C' if not is_ncr else 'B'}{row}")
        recov = _cell_val(ws, f"{g_col}{row}", 1)
        unit = ws[f"{'D' if not is_ncr else 'C'}{row}"].value

        if unit == "kg":
            cost = rate * cons
        else:  # percentage
            cost = rate * cons / 100

        eff = cost / recov if recov != 0 else 0

        ws[f"{f_col}{row}"] = round(cost, 2)
        ws[f"{h_col}{row}"] = round(eff, 2)
        ws[f"{t_col}{row}"] = round(eff, 2)
        total_cost += eff

    # --- Rows 12-14: Operating costs ---
    # Power (row 12)
    e12 = _cell_val(ws, f"{r_col}12")
    c12 = _cell_val(ws, f"{'C' if not is_ncr else 'B'}12")
    power = e12 * c12
    ws[f"{f_col}12"] = round(power, 2)
    ws[f"{t_col}12"] = round(power, 2)
    total_cost += power

    # Stores (row 13)
    cons_col = "C" if not is_ncr else "B"
    e13 = _cell_val(ws, f"{r_col}13")
    c13 = _cell_val(ws, f"{cons_col}13")
    f13 = e13 * c13
    ws[f"{f_col}13"] = round(f13, 2)
    i13 = _cell_val(ws, f"{t_col}13")
    if i13 == 0:
        i13 = f13
    # Write computed value (replace any formula)
    ws[f"{t_col}13"] = round(i13, 2)

    # Manpower (row 14)
    e14 = _cell_val(ws, f"{r_col}14")
    ws[f"{f_col}14"] = round(e14, 2)
    i14 = _cell_val(ws, f"{t_col}14")
    if i14 == 0:
        i14 = e14
    ws[f"{t_col}14"] = round(i14, 2)

    total_cost += i13 + i14

    # --- Row 15: Billet Cost ---
    ws[f"{t_col}15"] = round(total_cost, 2)
    billet_cost = total_cost

    # --- Row 16: Market price of Billet (already set as input) ---
    billet_mkt = _cell_val(ws, f"{t_col}16")

    # --- Row 17: Nett Margin ---
    ws[f"{t_col}17"] = round(billet_mkt - billet_cost, 2)

    # --- Section 2: Rolling Mill (rows 21-32) ---
    # Row 21: Cost of Billet
    ws["F21"] = round(billet_cost, 2)

    # Row 22: Scale loss
    c22 = _cell_val(ws, "C22") if not is_ncr else _cell_val(ws, "C22")
    f22 = billet_cost * c22 / 100
    ws["E22"] = round(billet_cost, 2)
    ws["F22"] = round(f22, 2)

    # Rows 23-26: Power, Stores, Manpower, Fuel
    f23 = _cell_val(ws, "E23") * _cell_val(ws, "C23")
    f24 = _cell_val(ws, "E24") * _cell_val(ws, "C24")
    f25 = _cell_val(ws, "E25") * _cell_val(ws, "C25")
    f26 = _cell_val(ws, "E26") * _cell_val(ws, "C26")
    ws["F23"] = round(f23, 2)
    ws["F24"] = round(f24, 2)
    ws["F25"] = round(f25, 2)
    ws["F26"] = round(f26, 2)

    # Row 27: Cutting loss
    f21_to_f26_sum = billet_cost + f22 + f23 + f24 + f25 + f26
    c27 = _cell_val(ws, "C27")
    f27 = f21_to_f26_sum * c27 / 100
    ws["E27"] = round(f21_to_f26_sum, 2)
    ws["F27"] = round(f27, 2)

    # Row 28: Rolling cost = SUM(F22:F27)
    f28 = f22 + f23 + f24 + f25 + f26 + f27
    ws["F28"] = round(f28, 2)

    # Row 29: Yield loss
    f21_to_f27_sum = f21_to_f26_sum + f27
    c29 = _cell_val(ws, "C29")
    f29 = f21_to_f27_sum * c29 / 100
    ws["E29"] = round(f21_to_f27_sum, 2)
    ws["F29"] = round(f29, 2)

    # Rows 30-31: Scrap recovery, Short length (already have values)
    f30 = _cell_val(ws, "E30") * _cell_val(ws, "C30")
    f31 = _cell_val(ws, "E31") * _cell_val(ws, "C31")
    ws["F30"] = round(f30, 2)
    ws["F31"] = round(f31, 2)

    # Row 32: Interest + Depreciation (hardcoded)
    f32 = _cell_val(ws, "F32")

    # Row 33: Total cost = F28 + Billet Cost + F32
    total_tmt_cost = f28 + billet_cost + f32
    ws["F33"] = round(total_tmt_cost, 2)

    # Row 34: Market price TMT (already set as input)
    tmt_mkt = _cell_val(ws, "F34")

    # Row 35: Margin = Market TMT - Total cost
    ws["F35"] = round(tmt_mkt - total_tmt_cost, 2)

    return billet_cost, total_tmt_cost


def update_workbook(wb, args):
    """Apply price updates and compute all values (no formulas left)."""
    updates_applied = []

    # --- Raipur Tab: Set inputs ---
    ws_r = wb["Raipur"]

    if args.pallet_dri is not None:
        ws_r["E7"] = args.pallet_dri
        updates_applied.append(f"Raipur E7 (Pallet DRI): {args.pallet_dri}")
    if args.pig_iron is not None:
        ws_r["E8"] = args.pig_iron
        updates_applied.append(f"Raipur E8 (Pig Iron): {args.pig_iron}")
    if args.scrap_raipur is not None:
        ws_r["E9"] = args.scrap_raipur
        updates_applied.append(f"Raipur E9 (Scrap): {args.scrap_raipur}")
    if args.silico_mn is not None:
        ws_r["E10"] = args.silico_mn
        updates_applied.append(f"Raipur E10 (Silico Mn): {args.silico_mn}")
    if args.iron_ore_dri is not None:
        ws_r["E11"] = args.iron_ore_dri
        updates_applied.append(f"Raipur E11 (Iron Ore DRI): {args.iron_ore_dri}")
    if args.report_date is not None:
        date_val = datetime.strptime(args.report_date, "%Y-%m-%d")
        ws_r["C15"] = date_val
        updates_applied.append(f"Raipur C15 (Date): {args.report_date}")
    if args.billet_raipur is not None:
        ws_r["I16"] = args.billet_raipur
        updates_applied.append(f"Raipur I16 (Billet Mkt): {args.billet_raipur}")
    if args.tmt_raipur is not None:
        ws_r["F34"] = args.tmt_raipur
        updates_applied.append(f"Raipur F34 (TMT Mkt): {args.tmt_raipur}")

    # Compute all Raipur formula cells
    _compute_tab(ws_r, {}, is_ncr=False)

    # --- NCR Tab: Set inputs (write computed values, not formulas) ---
    ws_n = wb["NCR"]

    # NCR rates derived from Raipur
    ws_n["D7"] = _cell_val(ws_r, "E7") + 3100
    ws_n["D8"] = _cell_val(ws_r, "E8") + 3100
    ws_n["D10"] = _cell_val(ws_r, "E10")
    ws_n["D11"] = _cell_val(ws_r, "E11") + 3100

    if args.scrap_mandi is not None:
        ws_n["D9"] = args.scrap_mandi - 500
        updates_applied.append(f"NCR D9 (Scrap): {args.scrap_mandi} - 500 = {args.scrap_mandi - 500}")
    if args.billet_mandi is not None:
        ws_n["H16"] = args.billet_mandi - 500
        updates_applied.append(f"NCR H16 (Billet Mkt): {args.billet_mandi} - 500 = {args.billet_mandi - 500}")
    if args.tmt_ncr is not None:
        ws_n["F34"] = args.tmt_ncr
        updates_applied.append(f"NCR F34 (TMT Mkt): {args.tmt_ncr}")

    # NCR date and label from Raipur
    ws_n["B15"] = ws_r["B15"].value
    ws_n["C15"] = ws_r["C15"].value

    # Compute all NCR formula cells
    _compute_tab(ws_n, {}, is_ncr=True)

    return updates_applied


def main():
    parser = argparse.ArgumentParser(
        description="Update costing Excel file with new prices"
    )
    parser.add_argument("excel_path", help="Path to the costing Excel file")
    parser.add_argument(
        "--output", "-o",
        help="Output file path (default: output/<filename>)",
        default=None,
    )

    # Raipur tab
    parser.add_argument("--pallet-dri", type=float, default=None)
    parser.add_argument("--pig-iron", type=float, default=None)
    parser.add_argument("--scrap-raipur", type=float, default=None)
    parser.add_argument("--silico-mn", type=float, default=None)
    parser.add_argument("--iron-ore-dri", type=float, default=None)
    parser.add_argument("--report-date", type=str, default=None)
    parser.add_argument("--billet-raipur", type=float, default=None)
    parser.add_argument("--tmt-raipur", type=float, default=None)

    # NCR tab
    parser.add_argument("--scrap-mandi", type=int, default=None)
    parser.add_argument("--billet-mandi", type=int, default=None)
    parser.add_argument("--tmt-ncr", type=float, default=None)

    args = parser.parse_args()

    if not os.path.exists(args.excel_path):
        print(f"ERROR: File not found: {args.excel_path}", file=sys.stderr)
        sys.exit(1)

    # Determine output path (date-wise folder when report-date is given)
    if args.output:
        output_path = args.output
    else:
        if args.report_date:
            # Name format: YYYYMMDD_Costing TMT.xlsx
            date_compact = args.report_date.replace("-", "")
            out_name = f"{date_compact}_Costing TMT.xlsx"
            output_path = os.path.join("output", args.report_date, out_name)
        else:
            basename = os.path.basename(args.excel_path)
            output_path = os.path.join("output", basename)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    print(f"Loading: {args.excel_path}", file=sys.stderr)
    wb = openpyxl.load_workbook(args.excel_path)

    # Keep only Raipur and NCR tabs, remove all others
    keep_tabs = {"Raipur", "NCR"}
    for name in wb.sheetnames:
        if name not in keep_tabs:
            del wb[name]
    print(f"Tabs kept: {wb.sheetnames}", file=sys.stderr)

    updates = update_workbook(wb, args)

    if not updates:
        print("WARNING: No updates specified. Use --help for options.", file=sys.stderr)
        sys.exit(0)

    # Apply professional formatting
    if "Raipur" in wb.sheetnames:
        format_raipur(wb["Raipur"])
    if "NCR" in wb.sheetnames:
        format_ncr(wb["NCR"])

    wb.save(output_path)
    print(f"Saved to: {output_path}", file=sys.stderr)

    # Safety check: fail if any formulas leaked through
    _verify_no_formulas(output_path)

    print(f"\nApplied {len(updates)} updates:", file=sys.stderr)
    for u in updates:
        print(f"  - {u}", file=sys.stderr)

    # Compute margins and update cumulative change log
    log_path = os.path.join("output", "change_log.xlsx")
    if args.report_date:
        margins = compute_margins(wb, args)
        print(f"\nComputed margins:", file=sys.stderr)
        print(f"  Raipur Billet Nett Margin: {margins['raipur_billet']}", file=sys.stderr)
        print(f"  Raipur TMT Margin:         {margins['raipur_tmt']}", file=sys.stderr)
        print(f"  NCR Billet Nett Margin:     {margins['ncr_billet']}", file=sys.stderr)
        print(f"  NCR TMT Margin:             {margins['ncr_tmt']}", file=sys.stderr)
        update_change_log(args, margins)
        format_change_log(log_path)

        # Auto-commit and push to main
        _auto_push(output_path, log_path, args.report_date)

    sys.exit(0)


if __name__ == "__main__":
    main()
