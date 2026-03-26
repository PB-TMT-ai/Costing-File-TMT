#!/usr/bin/env python3
"""Apply professional formatting to costing output Excel files.

Usage:
    python tools/format_output.py <excel_path> [--in-place]

Formats both the costing sheets (Raipur, NCR) and the change log.
By default saves alongside the original with '_formatted' suffix.
Use --in-place to overwrite the original file.

Exit codes:
    0 — Success
    1 — Error
"""

import argparse
import os
import sys

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter


# ── Colour palette ──────────────────────────────────────────────
DARK_BLUE = "1F4E79"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
INPUT_BLUE = "BDD7EE"
HEADER_GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
SUMMARY_GOLD = "FFF2CC"
MARGIN_POS = "C6EFCE"  # green tint for positive margin
MARGIN_NEG = "FFC7CE"  # red tint for negative margin
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
BORDER_COLOR = "B4C6E7"
DARK_BORDER = "1F4E79"

# ── Reusable style components ───────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

thick_bottom = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="medium", color=DARK_BORDER),
)

header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
title_font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
section_font = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
label_font = Font(name="Calibri", size=11, bold=False, color="333333")
bold_label = Font(name="Calibri", size=11, bold=True, color="333333")
value_font = Font(name="Calibri", size=11, bold=False, color="333333")
input_font = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
margin_font = Font(name="Calibri", size=11, bold=True, color="333333")

fill_dark_blue = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
fill_med_blue = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
fill_light_blue = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
fill_input = PatternFill(start_color=INPUT_BLUE, end_color=INPUT_BLUE, fill_type="solid")
fill_header_green = PatternFill(start_color=HEADER_GREEN, end_color=HEADER_GREEN, fill_type="solid")
fill_light_green = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
fill_gold = PatternFill(start_color=SUMMARY_GOLD, end_color=SUMMARY_GOLD, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_gray = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
fill_margin_pos = PatternFill(start_color=MARGIN_POS, end_color=MARGIN_POS, fill_type="solid")
fill_margin_neg = PatternFill(start_color=MARGIN_NEG, end_color=MARGIN_NEG, fill_type="solid")

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

INR_FORMAT = '#,##0'
INR_DECIMAL = '#,##0.00'
PCT_FORMAT = '0.0%'


def _apply_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Apply styling to a single cell."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _apply_row(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None):
    """Apply styling to a range of cells in a row."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        _apply_cell(cell, font=font, fill=fill, alignment=alignment, border=border)


def _apply_range(ws, row_start, row_end, col_start, col_end, border=None, fill=None):
    """Apply border/fill to a rectangular range."""
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            if border:
                cell.border = border
            if fill and not cell.fill.fgColor.rgb or cell.fill.fgColor.rgb == "00000000":
                cell.fill = fill


def format_raipur(ws):
    """Format the Raipur costing sheet."""
    # ── Title row ──
    _apply_cell(ws["B2"], font=title_font, alignment=left)

    # ── Section 1 header: Steel Costing RAIPUR ──
    for c in range(2, 10):  # B-I
        cell = ws.cell(row=4, column=c)
        _apply_cell(cell, font=header_font, fill=fill_dark_blue, alignment=center, border=thin_border)

    # ── Column headers (row 5) ──
    for c in range(2, 10):  # B-I
        cell = ws.cell(row=5, column=c)
        _apply_cell(cell, font=header_font, fill=fill_med_blue, alignment=center, border=thin_border)

    # ── Data rows 7-14 (raw materials + operating costs) ──
    for r in range(7, 15):
        is_even = (r % 2 == 0)
        bg = fill_gray if is_even else fill_white
        for c in range(1, 10):  # A-I
            cell = ws.cell(row=r, column=c)
            _apply_cell(cell, border=thin_border, fill=bg, font=label_font)
            if c == 2:  # Item name
                cell.alignment = left
            elif c >= 5:  # numeric columns
                cell.alignment = right_align

        # Highlight input cells (Rate column E) with blue
        _apply_cell(ws.cell(row=r, column=5), fill=fill_input, font=input_font, alignment=right_align)

    # ── Row 6: empty spacer — apply border to keep grid clean ──
    _apply_row(ws, 6, 1, 9, border=thin_border, fill=fill_white)

    # ── Billet Cost summary (row 15) ──
    for c in range(2, 10):
        cell = ws.cell(row=15, column=c)
        _apply_cell(cell, font=bold_label, fill=fill_gold, alignment=center, border=thick_bottom)
    ws["B15"].alignment = left

    # ── Market price Billet (row 16) ──
    for c in range(2, 10):
        cell = ws.cell(row=16, column=c)
        _apply_cell(cell, font=bold_label, fill=fill_light_green, alignment=center, border=thin_border)
    _apply_cell(ws["I16"], fill=fill_input, font=input_font, alignment=right_align)
    ws["B16"].alignment = left

    # ── Nett Margin (row 17) ──
    for c in range(2, 10):
        cell = ws.cell(row=17, column=c)
        _apply_cell(cell, font=margin_font, fill=fill_gold, alignment=center, border=thick_bottom)
    ws["B17"].alignment = left

    # ── Section 2: Rolling Mill Cost ──
    # Section header (row 19)
    for c in range(2, 7):  # B-F
        cell = ws.cell(row=19, column=c)
        _apply_cell(cell, font=header_font, fill=fill_header_green, alignment=center, border=thin_border)

    # Column headers (row 20)
    for c in range(2, 7):
        cell = ws.cell(row=20, column=c)
        _apply_cell(cell, font=header_font, fill=fill_med_blue, alignment=center, border=thin_border)

    # Data rows 21-32
    for r in range(21, 33):
        is_even = (r % 2 == 0)
        bg = fill_gray if is_even else fill_white
        for c in range(2, 7):  # B-F
            cell = ws.cell(row=r, column=c)
            _apply_cell(cell, border=thin_border, fill=bg, font=label_font)
            if c == 2:
                cell.alignment = left
            elif c >= 3:
                cell.alignment = right_align

    # Rolling cost subtotal (row 28)
    for c in range(2, 7):
        _apply_cell(ws.cell(row=28, column=c), font=bold_label, fill=fill_light_blue, border=thin_border)

    # ── Total cost (row 33) ──
    for c in range(2, 7):
        cell = ws.cell(row=33, column=c)
        _apply_cell(cell, font=bold_label, fill=fill_gold, alignment=center, border=thick_bottom)
    ws["B33"].alignment = left

    # ── Market price TMT (row 34) ──
    for c in range(2, 7):
        cell = ws.cell(row=34, column=c)
        _apply_cell(cell, font=bold_label, fill=fill_light_green, alignment=center, border=thin_border)
    _apply_cell(ws["F34"], fill=fill_input, font=input_font, alignment=right_align)
    ws["B34"].alignment = left

    # ── Margin for mfr (row 35) ──
    for c in range(2, 7):
        cell = ws.cell(row=35, column=c)
        _apply_cell(cell, font=margin_font, fill=fill_gold, alignment=center, border=thick_bottom)
    ws["B35"].alignment = left

    # ── Number formats ──
    for r in range(7, 15):
        for col in ["E", "F", "H", "I"]:
            cell = ws[f"{col}{r}"]
            if cell.value is not None:
                cell.number_format = INR_FORMAT
        ws[f"G{r}"].number_format = "0.00"

    for col in ["I15", "I16", "I17"]:
        ws[col].number_format = INR_FORMAT

    for r in range(21, 33):
        for col in ["E", "F"]:
            cell = ws[f"{col}{r}"]
            if cell.value is not None:
                cell.number_format = INR_FORMAT

    for col in ["F33", "F34", "F35"]:
        ws[col].number_format = INR_FORMAT

    # ── Column widths ──
    widths = {"A": 6, "B": 28, "C": 14, "D": 8, "E": 14, "F": 16, "G": 18, "H": 16, "I": 16}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # ── Freeze panes below headers ──
    ws.freeze_panes = "B6"

    # ── Print setup ──
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"


def format_ncr(ws):
    """Format the NCR costing sheet."""
    # NCR uses columns A-H instead of B-I

    # ── Title row ──
    _apply_cell(ws["A2"], font=title_font, alignment=left)

    # ── Section 1 header: Steel Costing NCR ──
    for c in range(1, 9):  # A-H
        cell = ws.cell(row=4, column=c)
        _apply_cell(cell, font=header_font, fill=fill_dark_blue, alignment=center, border=thin_border)

    # ── Column headers (row 5) ──
    for c in range(1, 9):  # A-H
        cell = ws.cell(row=5, column=c)
        _apply_cell(cell, font=header_font, fill=fill_med_blue, alignment=center, border=thin_border)

    # ── Data rows 7-14 ──
    for r in range(7, 15):
        is_even = (r % 2 == 0)
        bg = fill_gray if is_even else fill_white
        for c in range(1, 9):  # A-H
            cell = ws.cell(row=r, column=c)
            _apply_cell(cell, border=thin_border, fill=bg, font=label_font)
            if c == 1:  # Item name
                cell.alignment = left
            elif c >= 4:  # numeric columns
                cell.alignment = right_align

        # Highlight input cells (Rate column D) — only D9 is user input
        if r == 9:
            _apply_cell(ws.cell(row=r, column=4), fill=fill_input, font=input_font, alignment=right_align)

    # ── Row 6 spacer ──
    _apply_row(ws, 6, 1, 8, border=thin_border, fill=fill_white)

    # ── Billet Cost (row 15) ──
    for c in range(1, 9):
        cell = ws.cell(row=15, column=c)
        _apply_cell(cell, font=bold_label, fill=fill_gold, alignment=center, border=thick_bottom)
    ws["A15"].alignment = left

    # ── Market price Billet (row 16) ──
    for c in range(1, 9):
        cell = ws.cell(row=16, column=c)
        _apply_cell(cell, font=bold_label, fill=fill_light_green, alignment=center, border=thin_border)
    _apply_cell(ws["H16"], fill=fill_input, font=input_font, alignment=right_align)
    ws["A16"].alignment = left

    # ── Nett Margin (row 17) ──
    for c in range(1, 9):
        cell = ws.cell(row=17, column=c)
        _apply_cell(cell, font=margin_font, fill=fill_gold, alignment=center, border=thick_bottom)
    ws["A17"].alignment = left

    # ── Section 2: Rolling Mill Cost ──
    for c in range(1, 7):  # A-F (NCR rolling mill uses A-F based on structure)
        cell = ws.cell(row=19, column=c)
        _apply_cell(cell, font=header_font, fill=fill_header_green, alignment=center, border=thin_border)

    for c in range(1, 7):
        cell = ws.cell(row=20, column=c)
        _apply_cell(cell, font=header_font, fill=fill_med_blue, alignment=center, border=thin_border)

    for r in range(21, 33):
        is_even = (r % 2 == 0)
        bg = fill_gray if is_even else fill_white
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            _apply_cell(cell, border=thin_border, fill=bg, font=label_font)
            if c <= 2:
                cell.alignment = left
            elif c >= 3:
                cell.alignment = right_align

    for c in range(1, 7):
        _apply_cell(ws.cell(row=28, column=c), font=bold_label, fill=fill_light_blue, border=thin_border)

    # ── Total cost (row 33) ──
    for c in range(1, 7):
        cell = ws.cell(row=33, column=c)
        _apply_cell(cell, font=bold_label, fill=fill_gold, alignment=center, border=thick_bottom)
    ws["A33"].alignment = left

    # ── Market price TMT (row 34) ──
    for c in range(1, 7):
        cell = ws.cell(row=34, column=c)
        _apply_cell(cell, font=bold_label, fill=fill_light_green, alignment=center, border=thin_border)
    _apply_cell(ws["F34"], fill=fill_input, font=input_font, alignment=right_align)
    ws["A34"].alignment = left

    # ── Margin (row 35) ──
    for c in range(1, 7):
        cell = ws.cell(row=35, column=c)
        _apply_cell(cell, font=margin_font, fill=fill_gold, alignment=center, border=thick_bottom)
    ws["A35"].alignment = left

    # ── Number formats ──
    for r in range(7, 15):
        for col in ["D", "E", "G", "H"]:
            cell = ws[f"{col}{r}"]
            if cell.value is not None:
                cell.number_format = INR_FORMAT
        ws[f"F{r}"].number_format = "0.00"

    for col in ["H15", "H16", "H17"]:
        ws[col].number_format = INR_FORMAT

    for r in range(21, 33):
        for col in ["E", "F"]:
            cell = ws[f"{col}{r}"]
            if cell.value is not None:
                cell.number_format = INR_FORMAT

    for col in ["F33", "F34", "F35"]:
        ws[col].number_format = INR_FORMAT

    # ── Column widths ──
    widths = {"A": 28, "B": 16, "C": 14, "D": 14, "E": 14, "F": 18, "G": 16, "H": 16}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A6"

    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"


def format_change_log(log_path):
    """Format the change log Excel file."""
    if not os.path.exists(log_path):
        print(f"Change log not found: {log_path}", file=sys.stderr)
        return

    wb = openpyxl.load_workbook(log_path)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # ── Row 1: Date headers ──
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        _apply_cell(cell, font=header_font, fill=fill_dark_blue, alignment=center, border=thin_border)

    # ── Row 2: Raipur/NCR sub-headers ──
    for c in range(1, max_col + 1):
        cell = ws.cell(row=2, column=c)
        if c == 1:
            _apply_cell(cell, font=header_font, fill=fill_med_blue, alignment=center, border=thin_border)
        else:
            is_raipur = (c % 2 == 0)  # Raipur cols are even (2, 4, 6...)
            fill = fill_med_blue if is_raipur else PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            _apply_cell(cell, font=header_font, fill=fill, alignment=center, border=thin_border)

    # ── Item label column (A) ──
    for r in range(3, max_row + 1):
        cell = ws.cell(row=r, column=1)
        val = str(cell.value or "")
        if "Margin" in val or "Nett" in val:
            _apply_cell(cell, font=bold_label, fill=fill_gold, alignment=left, border=thin_border)
        elif "Market" in val:
            _apply_cell(cell, font=bold_label, fill=fill_light_green, alignment=left, border=thin_border)
        elif val == "Date":
            _apply_cell(cell, font=bold_label, fill=fill_light_blue, alignment=left, border=thin_border)
        else:
            _apply_cell(cell, font=label_font, fill=fill_white, alignment=left, border=thin_border)

    # ── Data cells ──
    for r in range(3, max_row + 1):
        item = str(ws.cell(row=r, column=1).value or "")
        for c in range(2, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value

            if "Margin" in item or "Nett" in item:
                # Colour margin cells based on positive/negative
                if isinstance(val, (int, float)):
                    bg = fill_margin_pos if val >= 0 else fill_margin_neg
                    _apply_cell(cell, font=margin_font, fill=bg, alignment=right_align,
                                border=thin_border, number_format=INR_DECIMAL)
                else:
                    _apply_cell(cell, font=label_font, fill=fill_gold, alignment=center, border=thin_border)
            elif "Market" in item:
                _apply_cell(cell, font=bold_label, fill=fill_light_green, alignment=right_align,
                            border=thin_border, number_format=INR_FORMAT)
            elif item == "Date":
                _apply_cell(cell, font=label_font, fill=fill_light_blue, alignment=center, border=thin_border)
            elif val == "-":
                _apply_cell(cell, font=label_font, fill=fill_gray, alignment=center, border=thin_border)
            elif isinstance(val, (int, float)):
                is_even = (r % 2 == 0)
                bg = fill_gray if is_even else fill_white
                _apply_cell(cell, font=value_font, fill=bg, alignment=right_align,
                            border=thin_border, number_format=INR_FORMAT)
            else:
                is_even = (r % 2 == 0)
                bg = fill_gray if is_even else fill_white
                _apply_cell(cell, font=value_font, fill=bg, alignment=center, border=thin_border)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 22
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ── Freeze panes ──
    ws.freeze_panes = "B3"

    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"

    wb.save(log_path)
    print(f"Change log formatted: {log_path}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(description="Format costing output Excel files")
    parser.add_argument("excel_path", help="Path to the costing Excel file to format")
    parser.add_argument(
        "--in-place", action="store_true",
        help="Overwrite the original file (default: save as *_formatted.xlsx)",
    )
    args = parser.parse_args()

    if not os.path.exists(args.excel_path):
        print(f"ERROR: File not found: {args.excel_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.excel_path)

    # Format Raipur tab
    if "Raipur" in wb.sheetnames:
        format_raipur(wb["Raipur"])
        print("Formatted: Raipur tab", file=sys.stderr)

    # Format NCR tab
    if "NCR" in wb.sheetnames:
        format_ncr(wb["NCR"])
        print("Formatted: NCR tab", file=sys.stderr)

    # Save
    if args.in_place:
        output_path = args.excel_path
    else:
        base, ext = os.path.splitext(args.excel_path)
        output_path = f"{base}_formatted{ext}"

    wb.save(output_path)
    print(f"Saved to: {output_path}", file=sys.stderr)

    # Also format change log if it exists alongside
    log_path = os.path.join(os.path.dirname(os.path.dirname(args.excel_path)), "change_log.xlsx")
    if not os.path.exists(log_path):
        log_path = os.path.join("output", "change_log.xlsx")
    if os.path.exists(log_path):
        format_change_log(log_path)

    sys.exit(0)


if __name__ == "__main__":
    main()
